from collections import deque
import openpyxl
from openpyxl.utils import get_column_letter
from config import ITEM_CATALOGUE


LOG_BUF   = deque(maxlen=60)
log_dirty = False

def tulis_log(pesan):
    global log_dirty
    LOG_BUF.append(f"• {pesan}")
    log_dirty = True

def get_log_dirty():
    return log_dirty

def set_log_dirty(val):
    global log_dirty
    log_dirty = val

def get_log_buf():
    return LOG_BUF

excel_transaction_counter = 0

def add_data_to_excel(npc, get_date_str_func, get_jam_str_func):
    """Write one sales row to Thrift_Sales_Dataset.xlsx when NPC pays at cashier."""
    global excel_transaction_counter

    item_name = npc.item_name
    if item_name not in ITEM_CATALOGUE:
        return

    category, price_per_unit = ITEM_CATALOGUE[item_name]
    quantity    = npc.quantity
    total_price = price_per_unit * quantity

    excel_transaction_counter += 1
    transaction_id = excel_transaction_counter

    customer_id = npc.nama
    gender      = "Female" if npc.gender == "female" else "Male"

    date_str    = get_date_str_func()
    time_str    = get_jam_str_func()
    combined_dt = f"{date_str} [{time_str}]"

    file_name = "Thrift_Sales_Dataset.xlsx"
    try:
        wb    = openpyxl.load_workbook(file_name)
        sheet = wb.active
    except Exception:
        wb    = openpyxl.Workbook()
        sheet = wb.active
        headers = ["Transaction_id", "Date", "Customer_id", "Gender",
                   "Product_category", "Product_name", "Quantity",
                   "Price_per_unit", "Total"]
        for col_idx, h in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_idx, value=h)

    new_row = sheet.max_row + 1
    sheet.cell(row=new_row, column=1, value=transaction_id)
    sheet.cell(row=new_row, column=2, value=combined_dt)
    sheet.cell(row=new_row, column=3, value=customer_id)
    sheet.cell(row=new_row, column=4, value=gender)
    sheet.cell(row=new_row, column=5, value=category)
    sheet.cell(row=new_row, column=6, value=item_name)
    sheet.cell(row=new_row, column=7, value=quantity)
    sheet.cell(row=new_row, column=8, value=price_per_unit)
    sheet.cell(row=new_row, column=9, value=total_price)

    for col in sheet.columns:
        max_len = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col)
        col_letter = get_column_letter(col[0].column)
        sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    try:
        wb.save(file_name)
        tulis_log(f"[Excel] #{transaction_id} {customer_id} → {item_name} x{quantity} = ${total_price}")
    except Exception as e:
        tulis_log(f"[Excel] Save error: {e}")
